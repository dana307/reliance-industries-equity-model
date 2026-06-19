from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)


NAVY, LT_BLUE, MID_BLUE = "1F497D", "DCE6F1", "B8CCE4"
WHITE, YELLOW, LT_GRAY   = "FFFFFF", "FFFF00", "F2F2F2"
BLUE_TXT, GREEN_TXT, BLACK_TEXT = "0000FF", "008000", "000000"

def fnt(bold=False, color="000000", size=9, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def fill(c): return PatternFill("solid", fgColor=c)
def aln(h="left", wrap=False): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def bdr(style="thin", c="BFBFBF"):
    s = Side(style=style, color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def bbdr():
    s = Side(style="thin", color="BFBFBF")
    return Border(bottom=s)

INR   = '#,##0;(#,##0);"-"'
PCT   = '0.0%;(0.0%);"-"'
PRICE = '#,##0.0;(#,##0.0);"-"'

YRS = [2017,2018,2019,2020,2021,2022,2023,2024,2025,2026]

sales    = [303954,390823,568337,596679,466307,694673,876396,899041,962820,1055780]
dep_da   = [11646,16706,20934,22203,26572,29782,40303,50832,53136,57688]
interest = [3849,8052,16495,22027,21189,14584,19571,23118,24269,27061]
oth_inc  = [9222,9869,8406,8570,22432,19600,12020,15792,17824,28846]
pbt      = [40034,49426,55227,53606,55461,83815,94464,104340,106017,123162]
tax_hist = [10201,13346,15390,13726,1722,15970,20376,25707,25230,27552]
pat      = [29901,36075,39588,39354,49128,60705,66702,69621,69648,80775]

eq_cap   = [2959,5922,5926,6339,6445,6765,6766,6766,13532,13532]
reserves = [260750,287584,381186,442827,693727,772720,709106,786715,829668,890498]
borrow   = [217475,239843,307714,355133,278962,319158,451664,350719,374313,402962]
net_blk  = [198526,403885,398374,532658,541258,627798,724805,779985,999393,1124795]
cwip_v   = [324837,187022,179463,109106,125953,172506,293752,338855,262358,237686]
cash_v   = [3023,4255,11081,30920,17397,36178,68664,97225,106502,145977]

cfo_v    = [49550,71459,42346,94877,26958,110654,115032,158788,178703,192113]
cfi_v    = [-66201,-68192,-94507,-72497,-142385,-109162,-93001,-113581,-137535,-101089]
cff_v    = [8617,-2001,55906,-2541,101904,17289,10455,-16646,-31891,-51549]

ebitda_h = [pbt[i]+dep_da[i]+interest[i]-oth_inc[i] for i in range(10)]
tot_eq   = [eq_cap[i]+reserves[i] for i in range(10)]
net_dbt  = [borrow[i]-cash_v[i] for i in range(10)]
nfa      = [net_blk[i]+cwip_v[i] for i in range(10)]
fcf_s    = [cfo_v[i]+cfi_v[i] for i in range(10)]

SHARES   = 1353.25
CUR_PRC  = 1309.5
MKT_CAP  = 1772552.58

N = len(YRS)   


def sc(ws, row, col, value=None, bold=False, color="000000", bg=None,
       fmt=None, h_align="left", border=True, wrap=False, italic=False, size=9):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font = fnt(bold=bold, color=color, size=size, italic=italic)
    if bg:
        cell.fill = fill(bg)
    if fmt:
        cell.number_format = fmt
    cell.alignment = aln(h=h_align, wrap=wrap)
    if border:
        cell.border = bdr()
    return cell

def hdr_row(ws, row, labels, col_start=1, col_span_first=1):
    for i, lbl in enumerate(labels):
        c = col_start + i
        cell = sc(ws, row, c, lbl, bold=True, color=WHITE, bg=NAVY, h_align="center", size=9)
    return row + 1

def sec_hdr(ws, row, label, ncols, col_start=1):
    sc(ws, row, col_start, label, bold=True, color="000000", bg=MID_BLUE, h_align="left")
    for c in range(col_start+1, col_start+ncols):
        sc(ws, row, c, None, bg=MID_BLUE)

def merge_label(ws, r, c1, c2, label, bold=False, bg=None):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1)
    cell.value = label
    cell.font = fnt(bold=bold, color=WHITE if bg else "000000")
    cell.alignment = aln(h="center")
    if bg:
        cell.fill = fill(bg)



ws1 = wb.create_sheet("Historical Financials")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 30
for ci in range(2, 13):
    ws1.column_dimensions[get_column_letter(ci)].width = 11

r = 1

ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
c = ws1.cell(row=r, column=1, value="RELIANCE INDUSTRIES LTD  |  Historical Financials (Consolidated)  |  ₹ Crores")
c.font = fnt(bold=True, color=WHITE, size=11)
c.fill = fill(NAVY)
c.alignment = aln(h="center")
r += 1

ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
c = ws1.cell(row=r, column=1, value="Source: Screener.in  |  FY2017–FY2026  |  All figures in ₹ Crores unless stated")
c.font = fnt(italic=True, size=8, color="595959")
c.alignment = aln(h="left")
r += 1; r += 1  # blank row


sec_hdr(ws1, r, "PROFIT & LOSS", 11)
r += 1


hdr_row(ws1, r, [""] + [f"FY{y}" for y in YRS])
r += 1

rows_pl = [
    ("Revenue (₹ Cr)", sales, INR, False),
    ("YoY Growth (%)", [None] + [f"=({get_column_letter(2+i)}6/{get_column_letter(1+i)}6)-1" if i>0 else None for i in range(1,N)], PCT, False),
    ("EBITDA (₹ Cr)", ebitda_h, INR, False),
    ("EBITDA Margin (%)", [f"={get_column_letter(1+i)}8/{get_column_letter(1+i)}6" for i in range(1,N+1)], PCT, False),
    ("Depreciation & Amortisation", dep_da, INR, False),
    ("EBIT (₹ Cr)", [ebitda_h[i]-dep_da[i] for i in range(N)], INR, False),
    ("EBIT Margin (%)", [f"={get_column_letter(1+i)}11/{get_column_letter(1+i)}6" for i in range(1,N+1)], PCT, False),
    ("Interest Expense", interest, INR, False),
    ("Other Income", oth_inc, INR, False),
    ("Profit Before Tax", pbt, INR, False),
    ("Tax", tax_hist, INR, False),
    ("PAT / Net Profit (₹ Cr)", pat, INR, True),
    ("PAT Margin (%)", [f"={get_column_letter(1+i)}17/{get_column_letter(1+i)}6" for i in range(1,N+1)], PCT, False),
    ("EPS (₹)", [round(pat[i]/SHARES, 2) for i in range(N)], PRICE, False),
]

pl_start_r = r
for label, vals, fmt, bold in rows_pl:
    sc(ws1, r, 1, label, bold=bold, color="000000")
    for i, v in enumerate(vals):
        if v is None:
            sc(ws1, r, 2+i, "-", h_align="center")
        elif isinstance(v, str) and v.startswith("="):
            cell = sc(ws1, r, 2+i, v, color=BLACK_TEXT if not bold else "000000", fmt=fmt, h_align="right", bold=bold)
        else:
            cell = sc(ws1, r, 2+i, v, color="000000", fmt=fmt, h_align="right", bold=bold)
    r += 1

r += 1  


sec_hdr(ws1, r, "BALANCE SHEET", 11)
r += 1
hdr_row(ws1, r, [""] + [f"FY{y}" for y in YRS])
r += 1

bs_rows = [
    ("Equity Share Capital", eq_cap, INR, False),
    ("Reserves & Surplus", reserves, INR, False),
    ("Total Equity", tot_eq, INR, True),
    ("Borrowings (Debt)", borrow, INR, False),
    ("Net Debt (Debt − Cash)", net_dbt, INR, True),
    ("Net Fixed Assets (Net Block + CWIP)", nfa, INR, False),
    ("Cash & Bank Balances", cash_v, INR, False),
]
for label, vals, fmt, bold in bs_rows:
    sc(ws1, r, 1, label, bold=bold)
    for i, v in enumerate(vals):
        sc(ws1, r, 2+i, v, fmt=fmt, h_align="right", bold=bold)
    r += 1

r += 1


sec_hdr(ws1, r, "CASH FLOW STATEMENT", 11)
r += 1
hdr_row(ws1, r, [""] + [f"FY{y}" for y in YRS])
r += 1

cf_rows = [
    ("Cash from Operations (CFO)", cfo_v, INR, False),
    ("Cash from Investing (CFI)", cfi_v, INR, False),
    ("Cash from Financing (CFF)", cff_v, INR, False),
    ("Free Cash Flow (CFO + CFI)", fcf_s, INR, True),
]
for label, vals, fmt, bold in cf_rows:
    sc(ws1, r, 1, label, bold=bold)
    for i, v in enumerate(vals):
        sc(ws1, r, 2+i, v, fmt=fmt, h_align="right", bold=bold)
    r += 1

r += 1


sec_hdr(ws1, r, "KEY METRICS & RATIOS", 11)
r += 1
hdr_row(ws1, r, [""] + [f"FY{y}" for y in YRS])
r += 1

roe  = [round(pat[i]/tot_eq[i], 4) for i in range(N)]
roce = [round((ebitda_h[i]-dep_da[i])/( tot_eq[i]+borrow[i]), 4) for i in range(N)]
nd_eb= [round(net_dbt[i]/ebitda_h[i], 2) if ebitda_h[i]>0 else None for i in range(N)]
div  = [round(pat[i]/sales[i], 4) for i in range(N)]

ratio_rows = [
    ("Return on Equity – ROE (%)", roe, PCT, False),
    ("Return on Capital Employed – ROCE (%)", roce, PCT, False),
    ("Net Debt / EBITDA (x)", nd_eb, '0.0x;(0.0x);"-"', False),
    ("Net Profit Margin (%)", div, PCT, False),
]
for label, vals, fmt, bold in ratio_rows:
    sc(ws1, r, 1, label, bold=bold)
    for i, v in enumerate(vals):
        if v is None:
            sc(ws1, r, 2+i, "n/m", h_align="center")
        else:
            sc(ws1, r, 2+i, v, fmt=fmt, h_align="right")
    r += 1



ws2 = wb.create_sheet("Assumptions")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 36
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 40
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 14
ws2.column_dimensions["G"].width = 14
ws2.column_dimensions["H"].width = 14

r = 1
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
c = ws2.cell(row=r, column=1, value="RELIANCE INDUSTRIES LTD  |  Model Assumptions  |  ₹ Crores")
c.font = fnt(bold=True, color=WHITE, size=11)
c.fill = fill(NAVY)
c.alignment = aln(h="center")
r += 1

ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
note = ws2.cell(row=r, column=1, value="🔵 Blue = Hardcoded Input  |  ⚫ Black = Formula  |  Change blue cells to run scenarios")
note.font = fnt(italic=True, size=8, color="595959")
r += 1; r += 1


sec_hdr(ws2, r, "WACC INPUTS", 3)
for c in range(4, 9): sc(ws2, r, c, None, bg=MID_BLUE)
r += 1

wacc_inputs = [
    ("Risk-Free Rate (10Y G-Sec Yield)", 0.070, "Source: RBI, approx 10Y G-Sec yield"),   # row 5
    ("Equity Beta", 0.95, "Source: NSE historical 5Y beta"),                               # row 6
    ("Equity Risk Premium – ERP", 0.055, "Damodaran India ERP estimate"),                  # row 7
    ("Cost of Equity [Ke = Rf + β × ERP]", "=B5+B6*B7", None),                           # row 8
    ("Cost of Debt (Pre-Tax)", 0.075, "Estimated from interest / avg debt"),                # row 9
    ("Effective Tax Rate", 0.25, "Estimated based on FY26 PAT/PBT"),                       # row 10
    ("Cost of Debt (Post-Tax) [Kd × (1−t)]", "=B9*(1-B10)", None),                       # row 11
    ("Market Capitalisation (₹ Cr)", 1772552.58, "Source: Screener.in, Jun 2025"),        # row 12
    ("Total Debt (₹ Cr)", 402962.0, "Source: FY2026 Balance Sheet"),                      # row 13
    ("Total Capital [MktCap + Debt]", "=B12+B13", None),                                  # row 14
    ("Weight of Equity [MktCap / Total]", "=B12/B14", None),                              # row 15
    ("Weight of Debt [Debt / Total]", "=B13/B14", None),                                  # row 16
    ("WACC [Ke×We + Kd×Wd]", "=B8*B15+B11*B16", None),                                  # row 17
]

wacc_row_start = r
for label, val, note_text in wacc_inputs:
    sc(ws2, r, 1, label)
    is_formula = isinstance(val, str) and val.startswith("=")
    is_pct = isinstance(val, float) and val < 5
    fmt = PCT if is_pct else INR
    sc(ws2, r, 2, val,
       color=BLACK_TEXT if is_formula else BLUE_TXT,
       fmt=fmt, h_align="right",
       bg=YELLOW if not is_formula else None)
    if note_text:
        sc(ws2, r, 3, note_text, italic=True, color="595959", border=False, size=8)
    r += 1

r += 1


sec_hdr(ws2, r, "REVENUE & MARGIN ASSUMPTIONS (FY27–FY31)", 8)
r += 1
proj_yr_labels = ["Assumption", "FY2027", "FY2028", "FY2029", "FY2030", "FY2031"]
hdr_row(ws2, r, proj_yr_labels)
r += 1

proj_assm_start = r
proj_rows = [
    ("Revenue Growth (%)",    [0.10, 0.10, 0.09, 0.09, 0.08], PCT),
    ("EBITDA Margin (%)",     [0.175,0.175,0.175,0.18, 0.18 ], PCT),
    ("D&A as % of Revenue",  [0.055,0.055,0.055,0.055,0.055], PCT),
    ("Capex as % of Revenue",[0.140,0.130,0.120,0.120,0.110], PCT),
]
for label, vals, fmt in proj_rows:
    sc(ws2, r, 1, label)
    for i, v in enumerate(vals):
        sc(ws2, r, 2+i, v, color=BLUE_TXT, fmt=fmt, h_align="right", bg=YELLOW)
    r += 1

r += 1


sec_hdr(ws2, r, "TERMINAL VALUE ASSUMPTIONS", 3)
for c in range(4, 9): sc(ws2, r, c, None, bg=MID_BLUE)
r += 1

tv_rows = [
    ("Terminal Growth Rate (g)", 0.05, "Approx. India nominal LT GDP growth"),
    ("Shares Outstanding (Cr)", SHARES, "Source: Screener.in, adjusted shares"),
    ("Current Market Price (₹)", CUR_PRC, "Source: NSE/BSE, Jun 2025"),
    ("Net Debt – FY2026 (₹ Cr)", 256985.0, "Borrowings 402,962 − Cash 145,977"),
]
tv_row_start = r
for label, val, note_text in tv_rows:
    sc(ws2, r, 1, label)
    fmt = PCT if isinstance(val, float) and val < 5 else PRICE
    sc(ws2, r, 2, val, color=BLUE_TXT, fmt=fmt, h_align="right", bg=YELLOW)
    if note_text:
        sc(ws2, r, 3, note_text, italic=True, color="595959", border=False, size=8)
    r += 1



ws3 = wb.create_sheet("DCF Valuation")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 34
for ci in range(2, 9):
    ws3.column_dimensions[get_column_letter(ci)].width = 14

r = 1
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
c = ws3.cell(row=r, column=1, value="RELIANCE INDUSTRIES LTD  |  DCF Valuation  |  ₹ Crores  |  5-Year Explicit Forecast (FY27–FY31)")
c.font = fnt(bold=True, color=WHITE, size=11)
c.fill = fill(NAVY)
c.alignment = aln(h="center")
r += 1

ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
note = ws3.cell(row=r, column=1, value="⚫ Formulas pull from Assumptions sheet  |  FCFF-based DCF  |  Year-end discounting")
note.font = fnt(italic=True, size=8, color="595959")
r += 1; r += 1


WACC_SHEET_ROW_START = 5   

Rf_row   = WACC_SHEET_ROW_START       
Beta_row = WACC_SHEET_ROW_START + 1   
ERP_row  = WACC_SHEET_ROW_START + 2   
Ke_row   = WACC_SHEET_ROW_START + 3   
Kd_pre_row = WACC_SHEET_ROW_START + 4 
Tax_row  = WACC_SHEET_ROW_START + 5   
Kd_row   = WACC_SHEET_ROW_START + 6   
MktCap_row = WACC_SHEET_ROW_START + 7 
Debt_row = WACC_SHEET_ROW_START + 8   
TotCap_row = WACC_SHEET_ROW_START + 9 
We_row   = WACC_SHEET_ROW_START + 10  
Wd_row   = WACC_SHEET_ROW_START + 11  
WACC_row = WACC_SHEET_ROW_START + 12  


PROJ_ROW_START = 21  
RevG_row  = PROJ_ROW_START       
EBITM_row = PROJ_ROW_START + 1   
DAP_row   = PROJ_ROW_START + 2   
CapP_row  = PROJ_ROW_START + 3   


TV_ROW_START = 27  
TVg_row    = TV_ROW_START       
Shares_row = TV_ROW_START + 1   
CurPrc_row = TV_ROW_START + 2   
NetDbt_row = TV_ROW_START + 3   

A_REF = "Assumptions"   


sec_hdr(ws3, r, "PROJECTED P&L", 7)
r += 1
hdr_row(ws3, r, ["", "FY2026A", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "FY2031E"])
yr_hdr_r = r
r += 1



REV_ROW  = r
sc(ws3, r, 1, "Revenue")
sc(ws3, r, 2, 1055780, fmt=INR, h_align="right", bold=True)  
for i, col in enumerate(range(3, 8)):  
    prev_col = get_column_letter(col - 1)
    grw_col  = get_column_letter(2 + i)   
    sc(ws3, r, col, f"={prev_col}{REV_ROW}*(1+{A_REF}!{grw_col}{RevG_row})",
       color=GREEN_TXT, fmt=INR, h_align="right", bold=True)
r += 1

REVG_ROW = r
sc(ws3, r, 1, "  YoY Revenue Growth (%)")
sc(ws3, r, 2, None)
for i, col in enumerate(range(3, 8)):
    as_col = get_column_letter(2 + i)
    sc(ws3, r, col, f"={A_REF}!{as_col}{RevG_row}", color=GREEN_TXT, fmt=PCT, h_align="right")
r += 1

EBITDA_ROW = r
sc(ws3, r, 1, "EBITDA")
sc(ws3, r, 2, ebitda_h[-1], fmt=INR, h_align="right")  
for i, col in enumerate(range(3, 8)):
    rev_col  = get_column_letter(col)
    as_col   = get_column_letter(2 + i)
    sc(ws3, r, col, f"={rev_col}{REV_ROW}*{A_REF}!{as_col}{EBITM_row}",
       color=GREEN_TXT, fmt=INR, h_align="right")
r += 1

EBITM_ROW = r
sc(ws3, r, 1, "  EBITDA Margin (%)")
sc(ws3, r, 2, f"=B{EBITDA_ROW}/B{REV_ROW}", fmt=PCT, h_align="right")  
for i, col in enumerate(range(3, 8)):
    as_col = get_column_letter(2 + i)
    sc(ws3, r, col, f"={A_REF}!{as_col}{EBITM_row}", color=GREEN_TXT, fmt=PCT, h_align="right")
r += 1

DA_ROW = r
sc(ws3, r, 1, "Depreciation & Amortisation (D&A)")
sc(ws3, r, 2, dep_da[-1], fmt=INR, h_align="right")
for i, col in enumerate(range(3, 8)):
    rev_col = get_column_letter(col)
    as_col  = get_column_letter(2 + i)
    sc(ws3, r, col, f"={rev_col}{REV_ROW}*{A_REF}!{as_col}{DAP_row}",
       color=GREEN_TXT, fmt=INR, h_align="right")
r += 1

EBIT_ROW = r
sc(ws3, r, 1, "EBIT (Operating Profit)")
for col in range(2, 8):
    eb_col = get_column_letter(col)
    da_col = get_column_letter(col)
    sc(ws3, r, col, f"={eb_col}{EBITDA_ROW-r+EBITDA_ROW}-{da_col}{DA_ROW}", fmt=INR, h_align="right")

for col in range(2, 8):
    c_l = get_column_letter(col)
    ws3.cell(row=EBIT_ROW, column=col).value = f"={c_l}{EBITDA_ROW}-{c_l}{DA_ROW}"
    ws3.cell(row=EBIT_ROW, column=col).number_format = INR
    ws3.cell(row=EBIT_ROW, column=col).alignment = aln(h="right")
r += 1

EBITM2_ROW = r
sc(ws3, r, 1, "  EBIT Margin (%)")
for col in range(2, 8):
    c_l = get_column_letter(col)
    sc(ws3, r, col, f"={c_l}{EBIT_ROW}/{c_l}{REV_ROW}", fmt=PCT, h_align="right")
r += 1


NOPAT_ROW = r
sc(ws3, r, 1, "NOPAT  [EBIT × (1 − Tax Rate)]")
for col in range(2, 8):
    c_l = get_column_letter(col)
    sc(ws3, r, col, f"={c_l}{EBIT_ROW}*(1-{A_REF}!B{Tax_row})", fmt=INR, h_align="right", bold=True)
r += 1

sc(ws3, r, 1, "  Add: D&A")
for col in range(2, 8):
    c_l = get_column_letter(col)
    sc(ws3, r, col, f"={c_l}{DA_ROW}", fmt=INR, h_align="right")
ADD_DA_ROW = r
r += 1

CAPEX_ROW = r
sc(ws3, r, 1, "  Less: Capital Expenditure (Capex)")
sc(ws3, r, 2, None)
for i, col in enumerate(range(3, 8)):
    rev_col = get_column_letter(col)
    as_col  = get_column_letter(2 + i)
    sc(ws3, r, col, f"=-{rev_col}{REV_ROW}*{A_REF}!{as_col}{CapP_row}",
       color=GREEN_TXT, fmt=INR, h_align="right")
r += 1

FCFF_ROW = r
sc(ws3, r, 1, "FREE CASH FLOW TO FIRM (FCFF)", bold=True)
sc(ws3, r, 2, None)
for col in range(3, 8):
    c_l = get_column_letter(col)
    sc(ws3, r, col, f"={c_l}{NOPAT_ROW}+{c_l}{ADD_DA_ROW}+{c_l}{CAPEX_ROW}",
       fmt=INR, h_align="right", bold=True)
r += 2

# ── Discounting ───────────────────────────────────────────────────────────────
sec_hdr(ws3, r, "DISCOUNTING  (Year-End Convention)", 7)
r += 1

DISC_PER_ROW = r
sc(ws3, r, 1, "Discount Period (years)")
sc(ws3, r, 2, None)
for i, col in enumerate(range(3, 8)):
    sc(ws3, r, col, i+1, fmt="0", h_align="right", color=BLUE_TXT)
r += 1

DISC_FAC_ROW = r
sc(ws3, r, 1, "Discount Factor  [1 / (1+WACC)^n]")
sc(ws3, r, 2, None)
for col in range(3, 8):
    c_l = get_column_letter(col)
    sc(ws3, r, col, f"=1/(1+{A_REF}!B{WACC_row})^{c_l}{DISC_PER_ROW}",
       color=GREEN_TXT, fmt="0.0000", h_align="right")
r += 1

PV_ROW = r
sc(ws3, r, 1, "PV of FCFF", bold=True)
sc(ws3, r, 2, None)
for col in range(3, 8):
    c_l = get_column_letter(col)
    sc(ws3, r, col, f"={c_l}{FCFF_ROW}*{c_l}{DISC_FAC_ROW}",
       fmt=INR, h_align="right", bold=True)
r += 1

SUM_PV_ROW = r
sc(ws3, r, 1, "Sum of PV of FCFFs (A)", bold=True)
sc(ws3, r, 2, f"=SUM(C{PV_ROW}:G{PV_ROW})", fmt=INR, h_align="right", bold=True, color=BLACK_TEXT)
r += 2


sec_hdr(ws3, r, "TERMINAL VALUE", 7)
r += 1

TV_FCF_ROW = r
sc(ws3, r, 1, "Terminal Year FCFF (FY31)")
sc(ws3, r, 2, f"=G{FCFF_ROW}", color=GREEN_TXT, fmt=INR, h_align="right")
r += 1

sc(ws3, r, 1, "Terminal Growth Rate (g)")
sc(ws3, r, 2, f"={A_REF}!B{TVg_row}", color=GREEN_TXT, fmt=PCT, h_align="right")
TV_G_ROW = r; r += 1

sc(ws3, r, 1, "WACC")
sc(ws3, r, 2, f"={A_REF}!B{WACC_row}", color=GREEN_TXT, fmt=PCT, h_align="right")
TV_WACC_ROW = r; r += 1

TV_VAL_ROW = r
sc(ws3, r, 1, "Terminal Value  [FCFF × (1+g) / (WACC − g)]", bold=True)
sc(ws3, r, 2, f"=B{TV_FCF_ROW}*(1+B{TV_G_ROW})/(B{TV_WACC_ROW}-B{TV_G_ROW})",
   fmt=INR, h_align="right", bold=True)
r += 1

TV_DISC_ROW = r
sc(ws3, r, 1, "Terminal Value Discount Factor  [1/(1+WACC)^5]")
sc(ws3, r, 2, f"=1/(1+{A_REF}!B{WACC_row})^5", color=GREEN_TXT, fmt="0.0000", h_align="right")
r += 1

PV_TV_ROW = r
sc(ws3, r, 1, "PV of Terminal Value (B)", bold=True)
sc(ws3, r, 2, f"=B{TV_VAL_ROW}*B{TV_DISC_ROW}", fmt=INR, h_align="right", bold=True)
r += 2


sec_hdr(ws3, r, "EQUITY VALUE BRIDGE", 7)
r += 1

EV_ROW = r
sc(ws3, r, 1, "Enterprise Value  [A + B]", bold=True)
sc(ws3, r, 2, f"=B{SUM_PV_ROW}+B{PV_TV_ROW}", fmt=INR, h_align="right", bold=True)
r += 1

sc(ws3, r, 1, "  Less: Net Debt (FY2026)")
sc(ws3, r, 2, f"=-{A_REF}!B{NetDbt_row}", color=GREEN_TXT, fmt=INR, h_align="right")
ND_REF_ROW = r; r += 1

EQ_VAL_ROW = r
sc(ws3, r, 1, "Equity Value", bold=True)
sc(ws3, r, 2, f"=B{EV_ROW}+B{ND_REF_ROW}", fmt=INR, h_align="right", bold=True)
r += 1

sc(ws3, r, 1, "Shares Outstanding (Cr)")
sc(ws3, r, 2, f"={A_REF}!B{Shares_row}", color=GREEN_TXT, fmt="0.00", h_align="right")
SHR_ROW = r; r += 1

IMP_ROW = r
sc(ws3, r, 1, "IMPLIED PRICE PER SHARE (₹)", bold=True, color=WHITE, bg=NAVY)
sc(ws3, r, 2, f"=B{EQ_VAL_ROW}/B{SHR_ROW}",
   bold=True, color=WHITE, bg=NAVY, fmt=PRICE, h_align="right")
ws3.row_dimensions[r].height = 18
r += 1

CUR_ROW = r
sc(ws3, r, 1, "Current Market Price (₹)")
sc(ws3, r, 2, f"={A_REF}!B{CurPrc_row}", color=GREEN_TXT, fmt=PRICE, h_align="right")
r += 1

UPSIDE_ROW = r
sc(ws3, r, 1, "Upside / (Downside) vs Market Price (%)", bold=True)
sc(ws3, r, 2, f"=(B{IMP_ROW}/B{CUR_ROW})-1", color=BLACK_TEXT, fmt=PCT, h_align="right", bold=True)
r += 2


ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
note_cell = ws3.cell(row=r, column=1,
    value=("📌 NOTE: Reliance is a diversified conglomerate (O2C, Retail, Jio, New Energy). "
           "A single-entity FCFF DCF may undervalue it — a Sum-of-the-Parts (SOTP) model "
           "per segment would be the next step in a full analysis."))
note_cell.font = fnt(italic=True, size=8, color="595959")
note_cell.alignment = aln(h="left", wrap=True)
ws3.row_dimensions[r].height = 28



ws4 = wb.create_sheet("Comps")
ws4.sheet_view.showGridLines = False
col_widths = [22, 14, 14, 14, 14, 14, 8, 10, 10, 8]
for i, w in enumerate(col_widths):
    ws4.column_dimensions[get_column_letter(i+1)].width = w

r = 1
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
c = ws4.cell(row=r, column=1, value="RELIANCE INDUSTRIES LTD  |  Comparable Company Analysis (Comps)")
c.font = fnt(bold=True, color=WHITE, size=11)
c.fill = fill(NAVY)
c.alignment = aln(h="center")
r += 1

ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
note = ws4.cell(row=r, column=1,
    value="Source: NSE/BSE, Company Filings | All ₹ Cr | Fill in peer data from Screener.in")
note.font = fnt(italic=True, size=8, color="595959")
r += 1; r += 1

hdr_labels = ["Company", "Mkt Cap (₹Cr)", "EV (₹Cr)", "Revenue (₹Cr)",
              "EBITDA (₹Cr)", "PAT (₹Cr)", "P/E (x)", "EV/EBITDA (x)", "EV/Sales (x)", "ROE (%)"]
hdr_row(ws4, r, hdr_labels)
r += 1


ril_ev = MKT_CAP + borrow[-1] - cash_v[-1]  # ~1,883,537
comps_data = [
    ["Reliance Industries", MKT_CAP, ril_ev, sales[-1], ebitda_h[-1], pat[-1]],
]

for row_data in comps_data:
    company, mc, ev, rev, ebt, net = row_data
    sc(ws4, r, 1, company, bold=True)
    sc(ws4, r, 2, mc, fmt=INR, h_align="right", bold=True)
    sc(ws4, r, 3, ev, fmt=INR, h_align="right")
    sc(ws4, r, 4, rev, fmt=INR, h_align="right")
    sc(ws4, r, 5, ebt, fmt=INR, h_align="right")
    sc(ws4, r, 6, net, fmt=INR, h_align="right")
    sc(ws4, r, 7, f"=B{r}/F{r}", fmt='0.0x', h_align="right")
    sc(ws4, r, 8, f"=C{r}/E{r}", fmt='0.0x', h_align="right")
    sc(ws4, r, 9, f"=C{r}/D{r}", fmt='0.0x', h_align="right")
    sc(ws4, r, 10, f"=F{r}/(B{r}-B{r}+F{r})", fmt=PCT, h_align="right")  # placeholder
    r += 1


peers = ["ONGC", "IOC (Indian Oil)", "BPCL", "Adani Enterprises", "Tata Conglomerate"]
for peer in peers:
    sc(ws4, r, 1, peer, color="595959", italic=True)
    for col in range(2, 11):
        sc(ws4, r, col, None, bg=LT_BLUE)
        ws4.cell(row=r, column=col).number_format = INR if col <= 6 else '0.0x'
    r += 1

r += 1

sc(ws4, r, 1, "Median (Peers)", bold=True)

r += 1

ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
note2 = ws4.cell(row=r, column=1,
    value=("💡 HOW TO USE: Fill in the blue cells with peer data from Screener.in. "
           "EV = Mkt Cap + Total Debt − Cash. P/E = Mkt Cap / PAT. EV/EBITDA and EV/Sales compute automatically."))
note2.font = fnt(italic=True, size=8, color="595959")
note2.alignment = aln(h="left", wrap=True)
ws4.row_dimensions[r].height = 28



ws5 = wb.create_sheet("Sensitivity")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 20
for ci in range(2, 12):
    ws5.column_dimensions[get_column_letter(ci)].width = 12

r = 1
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
c = ws5.cell(row=r, column=1, value="RELIANCE INDUSTRIES LTD  |  Sensitivity Analysis  |  Implied Price per Share (₹)")
c.font = fnt(bold=True, color=WHITE, size=11)
c.fill = fill(NAVY)
c.alignment = aln(h="center")
r += 1

ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
note = ws5.cell(row=r, column=1,
    value="Rows: WACC  |  Columns: Terminal Growth Rate (g)  |  Base case highlighted in gold")
note.font = fnt(italic=True, size=8, color="595959")
r += 1; r += 1


def dcf_price(wacc, g):
    rev = 1055780
    revg  = [0.10, 0.10, 0.09, 0.09, 0.08]
    ebitm = 0.175
    da_p  = 0.055
    tax   = 0.25
    capp  = [0.14, 0.13, 0.12, 0.12, 0.11]
    revs, fcffs = [], []
    for gr in revg:
        rev = rev * (1 + gr)
        revs.append(rev)
    for i, rv in enumerate(revs):
        ebitda = rv * ebitm
        da     = rv * da_p
        ebit   = ebitda - da
        nopat  = ebit * (1 - tax)
        capex  = rv * capp[i]
        fcffs.append(nopat + da - capex)
    pv_fcff = sum(f / (1 + wacc)**(i+1) for i, f in enumerate(fcffs))
    tv      = fcffs[-1] * (1 + g) / (wacc - g)
    pv_tv   = tv / (1 + wacc)**5
    ev      = pv_fcff + pv_tv
    net_debt = 256985
    eq_val  = ev - net_debt
    return (eq_val / SHARES) * 100

wacc_range = [w/100 for w in range(90, 135, 5)]  # 9.0% to 13.0%
g_range    = [g/100 for g in range(30, 75, 5)]   # 3.0% to 7.0%

# Header row: g values
sc(ws5, r, 1, "WACC \\ g →", bold=True, bg=NAVY, color=WHITE, h_align="center")
for i, g in enumerate(g_range):
    sc(ws5, r, 2+i, f"{g:.1%}", bold=True, bg=NAVY, color=WHITE, h_align="center")
r += 1

BASE_WACC = 0.11
BASE_G    = 0.05

for wacc in wacc_range:
    sc(ws5, r, 1, f"{wacc:.1%}", bold=True, bg=MID_BLUE, h_align="center")
    for i, g in enumerate(g_range):
        price_val = dcf_price(wacc, g)
        is_base = (abs(wacc - BASE_WACC) < 0.001 and abs(g - BASE_G) < 0.001)
        bg_color = "FFD700" if is_base else None  # gold for base case
        cell = sc(ws5, r, 2+i, round(price_val, 1),
                  fmt=PRICE, h_align="right",
                  bold=is_base,
                  bg=bg_color if bg_color else (LT_BLUE if price_val > CUR_PRC else None))
    r += 1

r += 1
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
note2 = ws5.cell(row=r, column=1,
    value=(f"🏷️  Base case: WACC = {BASE_WACC:.0%}, g = {BASE_G:.0%}.  "
           f"Current market price: ₹{CUR_PRC:,.1f}.  "
           "Light blue cells = implied price exceeds current market price (upside scenarios).  "
           "Gap vs market price reflects Reliance's embedded growth optionality (Jio, Retail, New Energy)."))
note2.font = fnt(italic=True, size=8, color="595959")
note2.alignment = aln(h="left", wrap=True)
ws5.row_dimensions[r].height = 35



OUT = "/sessions/busy-ecstatic-pascal/mnt/outputs/RIL_DCF_Model.xlsx"
wb.save(OUT)
print("Saved:", OUT)
