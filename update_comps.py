from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY, LT_BLUE, MID_BLUE = "1F497D", "DCE6F1", "B8CCE4"
WHITE, LT_GRAY = "FFFFFF", "F2F2F2"
BLUE_TXT, GREEN_TXT, BLACK_TEXT = "0000FF", "008000", "000000"

INR   = '#,##0;(#,##0);"-"'
PCT   = '0.0%;(0.0%);"-"'
MUL   = '0.0x'

def fnt(bold=False, color="000000", size=9, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def fill(c): return PatternFill("solid", fgColor=c)
def aln(h="left", wrap=False): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def side(style="thin", c="BFBFBF"): return Side(style=style, color=c)
def bdr(): s = side(); return Border(left=s, right=s, top=s, bottom=s)

def sc(ws, row, col, value=None, bold=False, color="000000", bg=None,
       fmt=None, h_align="left"):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font = fnt(bold=bold, color=color)
    if bg:
        cell.fill = fill(bg)
    elif bg == "":
        cell.fill = PatternFill(fill_type=None)
    if fmt:
        cell.number_format = fmt
    cell.alignment = aln(h=h_align)
    cell.border = bdr()

wb = load_workbook("RIL_DCF_Model.xlsx")
ws = wb["Comps"]



sc(ws, 5, 10, 0.0925, color=BLACK_TEXT, fmt=PCT, h_align="right")


peers = [
    
    ("Adani Enterprises",  263909,   347728,   97895,  14252,   8005, 0.0982),
    ("Tata Motors",        266787,   303327,  439695,  56138,  28149, 0.2360),
    ("Larsen & Toubro",    535675,   613279,  255734,  34427,  17673, 0.1590),
    ("Bharti Airtel",     1113302,  1308944,  172985,  85060,  37481, 0.2320),
    ("ITC",                355462,   321027,   75323,  25839,  20104, 0.2930),
]

for i, (name, mc, ev, rev, ebitda, pat, roe) in enumerate(peers):
    r = 6 + i
    sc(ws, r, 1, name, bold=False, color="000000")
    ws.cell(row=r, column=1).fill = PatternFill(fill_type=None)

    sc(ws, r, 2, mc,    color=BLUE_TXT, fmt=INR, h_align="right")
    sc(ws, r, 3, ev,    color=BLUE_TXT, fmt=INR, h_align="right")
    sc(ws, r, 4, rev,   color=BLUE_TXT, fmt=INR, h_align="right")
    sc(ws, r, 5, ebitda,color=BLUE_TXT, fmt=INR, h_align="right")
    sc(ws, r, 6, pat,   color=BLUE_TXT, fmt=INR, h_align="right")
    sc(ws, r, 7, f"=B{r}/F{r}", color=BLACK_TEXT, fmt=MUL, h_align="right")
    sc(ws, r, 8, f"=C{r}/E{r}", color=BLACK_TEXT, fmt=MUL, h_align="right")
    sc(ws, r, 9, f"=C{r}/D{r}", color=BLACK_TEXT, fmt=MUL, h_align="right")
    sc(ws, r, 10, roe,  color=BLUE_TXT, fmt=PCT, h_align="right")

    
    for col in [2, 3, 4, 5, 6, 10]:
        ws.cell(row=r, column=col).fill = PatternFill(fill_type=None)

sc(ws, 12, 1, "Median (Peers)", bold=True)
for col, fmt in [(2, INR), (3, INR), (4, INR), (5, INR), (6, INR),
                  (7, MUL), (8, MUL), (9, MUL), (10, PCT)]:
    col_letter = ["A","B","C","D","E","F","G","H","I","J"][col-1]
    sc(ws, 12, col, f"=MEDIAN({col_letter}6:{col_letter}10)",
       color=BLACK_TEXT, fmt=fmt, h_align="right")

wb.save("RIL_DCF_Model.xlsx")
print("Comps sheet updated.")
